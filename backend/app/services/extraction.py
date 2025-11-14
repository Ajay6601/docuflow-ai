import logging
from typing import Optional, Tuple
from io import BytesIO
import PyPDF2
import pytesseract
from PIL import Image
from docx import Document as DocxDocument
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExtractionService:
    """Service for extracting text from various document types."""
    
    @staticmethod
    def extract_from_pdf(file_data: bytes) -> Tuple[str, int, str]:
        """
        Extract text from PDF file.
        
        Returns:
            Tuple of (extracted_text, page_count, method_used)
        """
        try:
            pdf_file = BytesIO(file_data)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            page_count = len(pdf_reader.pages)
            
            extracted_text = ""
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    text = page.extract_text()
                    if text:
                        extracted_text += f"\n--- Page {page_num + 1} ---\n{text}"
                except Exception as e:
                    logger.warning(f"Error extracting text from page {page_num + 1}: {e}")
            
            # If no text extracted, it might be a scanned PDF - try OCR
            if not extracted_text.strip():
                logger.info("No text found in PDF, attempting OCR...")
                extracted_text = ExtractionService._ocr_pdf(file_data)
                method = "ocr"
            else:
                method = "pdf_text"
            
            return extracted_text.strip(), page_count, method
            
        except Exception as e:
            logger.error(f"Error extracting from PDF: {e}")
            raise
    
    @staticmethod
    def _ocr_pdf(file_data: bytes) -> str:
        """
        Perform OCR on PDF (for scanned documents).
        Note: This is a basic implementation. For production, use pdf2image + tesseract.
        """
        try:
            # For now, return a placeholder
            # In production, you'd use pdf2image to convert PDF pages to images
            # then run OCR on each image
            return "[OCR extraction from PDF requires pdf2image - placeholder]"
        except Exception as e:
            logger.error(f"Error during PDF OCR: {e}")
            return ""
    
    @staticmethod
    def extract_from_image(file_data: bytes) -> Tuple[str, str]:
        """
        Extract text from image using OCR.
        
        Returns:
            Tuple of (extracted_text, method_used)
        """
        try:
            image = Image.open(BytesIO(file_data))
            
            # Perform OCR
            extracted_text = pytesseract.image_to_string(image)
            
            return extracted_text.strip(), "ocr"
            
        except Exception as e:
            logger.error(f"Error extracting from image: {e}")
            raise
    
    @staticmethod
    def extract_from_docx(file_data: bytes) -> Tuple[str, str]:
        """
        Extract text from DOCX file.
        
        Returns:
            Tuple of (extracted_text, method_used)
        """
        try:
            doc = DocxDocument(BytesIO(file_data))
            
            extracted_text = ""
            for para in doc.paragraphs:
                extracted_text += para.text + "\n"
            
            return extracted_text.strip(), "docx"
            
        except Exception as e:
            logger.error(f"Error extracting from DOCX: {e}")
            raise
    
    @staticmethod
    def extract_from_xlsx(file_data: bytes) -> Tuple[str, str]:
        """
        Extract text from XLSX file.
        
        Returns:
            Tuple of (extracted_text, method_used)
        """
        try:
            workbook = load_workbook(BytesIO(file_data), data_only=True)
            
            extracted_text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                extracted_text += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        extracted_text += row_text + "\n"
            
            return extracted_text.strip(), "xlsx"
            
        except Exception as e:
            logger.error(f"Error extracting from XLSX: {e}")
            raise
    
    @staticmethod
    def extract_text(file_data: bytes, mime_type: str) -> Tuple[Optional[str], Optional[int], Optional[str], Optional[str]]:
        """
        Main extraction method that routes to appropriate extractor.
        
        Args:
            file_data: File content as bytes
            mime_type: MIME type of the file
            
        Returns:
            Tuple of (extracted_text, page_count, method_used, error_message)
        """
        try:
            if mime_type == "application/pdf":
                text, page_count, method = ExtractionService.extract_from_pdf(file_data)
                return text, page_count, method, None
            
            elif mime_type in ["image/png", "image/jpeg", "image/jpg"]:
                text, method = ExtractionService.extract_from_image(file_data)
                return text, None, method, None
            
            elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                text, method = ExtractionService.extract_from_docx(file_data)
                return text, None, method, None
            
            elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                text, method = ExtractionService.extract_from_xlsx(file_data)
                return text, None, method, None
            
            else:
                return None, None, None, f"Unsupported file type for extraction: {mime_type}"
                
        except Exception as e:
            error_msg = f"Extraction failed: {str(e)}"
            logger.error(error_msg)
            return None, None, None, error_msg


# Singleton instance
extraction_service = ExtractionService()